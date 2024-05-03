import openpyxl
#subject = ['English','English Gramer', 'Mathemetics']
subject = ['English','English Gramer', 'Mathemetics','Science', 'Moral Science', 'Computer', 'Gujarati','Hindi','social Science', 'Sanskrit','Drawing','P.T.']
subject_short=['eng','eng_gra','maths','sci','moral','comp','guj','hin','ss','san','draw','pt']
name_size=10
subject_size=9
font_size=8
reopen_date="13-06-2024"
issue_date="04-05-2024"
no_of_student=22
file_Name = "marks_7.xlsx"
total_working_day=232
 
# To open the workbook 
# workbook object is created
wb = openpyxl.load_workbook(file_Name)
#which sheet
ws = wb["Sheet1"]

from fpdf import FPDF
class PDF(FPDF):
    def header(self):
        # Rendering logo:
        #self.image("../docs/fpdf2-logo.png", 10, 8, 33)
        # Setting font: helvetica bold 15
        self.set_font("helvetica", "B", 15)
        # Moving cursor to the right:
        #self.cell(80)
        self.cell(250, 5, "Saint Mary School,Bagasara", align="C")
        # Performing a line break:
        self.ln(2)
        self.set_font("helvetica","I",11)
        self.cell(250,12,"Amreli Road, Bagasara - 365440, Dist. Amreli, Gujarat   Ph.:9428209955", align="C")
        self.ln(5)
        self.set_font("helvetica", "B", 13)
        self.cell(250, 15, "ANNUAL REPORT CARD: 2023-24", align="C")
        self.ln(7)

    
    def footer(self):
        # Position cursor at 1.5 cm from bottom:
        self.ln(3)
        self.set_y(-25)
        # Setting font: helvetica italic 8
        self.set_font("helvetica", "B", 12)
        self.cell(250, 12, f"School Reopens on : {reopen_date}", align="L")
        self.ln(5)
        self.cell(250, 15, f"Issue Date : {issue_date}                            Signature of Class Teacher:                                    Signature of Principal:", align="L")
        self.ln(5)
        self.cell(250,20, "Grading System:    91-100 : A+    81-90: A    71-80 : B+      61-70: B    51-60 : C+      41-50: C    31-40: D+", align="L")
        

        # Printing page number:
        #self.cell(0, 10, f"Page {self.page_no()}/{{nb}}", align="C")
    

# Instantiation of inherited class
def result_pdf(data,pdf):
    pdf.add_page()
    pdf.ln(5)
    with pdf.table() as table:
        pdf.set_font("helvetica", "B", name_size)
        row = table.row()
        row.cell("Name")
        row.cell(data["name"],colspan=5)

        row = table.row()
        row.cell("Standard:")
        row.cell(data['std'])
        row.cell("GR NO:")
        row.cell(data["grno"])
        row.cell("Roll No:")
        row.cell(data['roll_no'])

    pdf.set_font("helvetica", "B", subject_size)
    with pdf.table() as table:
        row = table.row()

        row = table.row()
        row.cell("Subject", rowspan=2,align="C")
        row.cell("Semester-1",align="C")
        row.cell("Semester-2",align="C")
        row.cell("Annual",colspan=2,align="C")
        
        row = table.row()
        row.cell("Marks(100)",align="C")
        row.cell("Marks(100)",align="C")
        row.cell("Marks(200)",align="C")
        row.cell("Grade(Sem1+Sem2)",align="C")
        sub_no=0
        for sub in subject:
            row = table.row()
            pdf.set_font("helvetica", "B", subject_size)
            row.cell(sub)
            sub_short=[]
            sub_short.append(subject_short[sub_no]+"_sem1")
            sub_short.append(subject_short[sub_no]+"_sem2")
            sub_short.append(subject_short[sub_no]+"_annual")
            sub_short.append(subject_short[sub_no]+"_grade")
            pdf.set_font("helvetica", "", font_size)
            
            for s in sub_short:
                row.cell(data[s],align="C")    
            '''
            row.cell("<<eng_sem1>>",align="C")
            row.cell("<<eng_sem2>>",align="C")
            row.cell("<<Annual>>",align="C")
            row.cell("<<grade1>>",align="C")
            '''
            sub_no = sub_no +1
        row = table.row()
        pdf.set_font("helvetica", "B", subject_size)
        row.cell("Obtained Total Marks & Overall Grade :",align="L",colspan=3)
        row.cell(data['total'],align="C")
        row.cell(data['grade'],align="C")
        row = table.row()
        row.cell("Obtained Percentage :",align="L",colspan=4)
        row.cell(data['per'],align="C")
    
        row = table.row()
        row.cell("Attendance :",align="L")
        row.cell("«Attend»",align="C")
        row.cell("Total Working Days :",align="C",colspan=2)
        row.cell(str(total_working_day),align="C")

        row = table.row()
        row.cell("Passed and Promoted to next Standard :",align="L",colspan=4)
        row.cell("",align="C")

        row = table.row()
        row.cell("Remarks :",align="L",colspan=5)

def fetchData(no):
    row_no=no+1
    student_data={}
    student_data["std"]  = ws["A"+str(row_no)].value
    student_data['roll_no'] = str(ws["B"+str(row_no)].value)
    student_data['grno'] = str(ws["C"+str(row_no)].value)
    student_data["name"] = ws["D"+str(row_no)].value
    #English Subject Detail
    student_data['eng_sem1'] = str(round(ws["F"+str(row_no)].value,1))
    student_data['eng_sem2'] = str(round(ws["G"+str(row_no)].value,1))
    student_data['eng_annual'] = str(round(ws["H"+str(row_no)].value,1))
    student_data['eng_grade'] = ws["I"+str(row_no)].value
    #English Grammer
    student_data['eng_gra_sem1'] = str(round(ws["j"+str(row_no)].value,1))
    student_data['eng_gra_sem2'] = str(round(ws["k"+str(row_no)].value,1))
    student_data['eng_gra_annual'] = str(round(ws["l"+str(row_no)].value,1))
    student_data['eng_gra_grade'] = ws["m"+str(row_no)].value
    #Maths
    student_data['maths_sem1'] = str(round(ws["n"+str(row_no)].value,1))
    student_data['maths_sem2'] = str(round(ws["o"+str(row_no)].value,1))
    student_data['maths_annual'] = str(round(ws["p"+str(row_no)].value,1))
    student_data['maths_grade'] = ws["q"+str(row_no)].value
    #Science
    student_data['sci_sem1'] = str(round(ws["r"+str(row_no)].value,1))
    student_data['sci_sem2'] = str(round(ws["s"+str(row_no)].value,1))
    student_data['sci_annual'] = str(round(ws["t"+str(row_no)].value,1))
    student_data['sci_grade'] = ws["u"+str(row_no)].value
    #Social Science
    student_data['ss_sem1'] = str(round(ws["v"+str(row_no)].value,1))
    student_data['ss_sem2'] = str(round(ws["w"+str(row_no)].value,1))
    student_data['ss_annual'] = str(round(ws["x"+str(row_no)].value,1))
    student_data['ss_grade'] = ws["y"+str(row_no)].value
    #Gujarati
    student_data['guj_sem1'] = str(round(ws["z"+str(row_no)].value,1))
    student_data['guj_sem2'] = str(round(ws["aa"+str(row_no)].value,1))
    student_data['guj_annual'] = str(round(ws["ab"+str(row_no)].value,1))
    student_data['guj_grade'] = ws["ac"+str(row_no)].value
    #Hindi
    student_data['hin_sem1'] = str(round(ws["ad"+str(row_no)].value,1))
    student_data['hin_sem2'] = str(round(ws["ae"+str(row_no)].value,1))
    student_data['hin_annual'] = str(round(ws["af"+str(row_no)].value,1))
    student_data['hin_grade'] = ws["ag"+str(row_no)].value
    #Sanskrit
    student_data['san_sem1'] = str(round(ws["ah"+str(row_no)].value,1))
    student_data['san_sem2'] = str(round(ws["ai"+str(row_no)].value,1))
    student_data['san_annual'] = str(round(ws["aj"+str(row_no)].value,1))
    student_data['san_grade'] = ws["ak"+str(row_no)].value
    #Computer
    student_data['comp_sem1'] = str(round(ws["al"+str(row_no)].value,1))
    student_data['comp_sem2'] = str(round(ws["am"+str(row_no)].value,1))
    student_data['comp_annual'] = str(round(ws["an"+str(row_no)].value,1))
    student_data['comp_grade'] = ws["ao"+str(row_no)].value
    #Moral Science
    student_data['moral_sem1'] = str(round(ws["ap"+str(row_no)].value,1))
    student_data['moral_sem2'] = str(round(ws["aq"+str(row_no)].value,1))
    student_data['moral_annual'] = str(round(ws["ar"+str(row_no)].value,1))
    student_data['moral_grade'] = ws["as"+str(row_no)].value
    #Drwing
    student_data['draw_sem1'] = ""
    student_data['draw_sem2'] = ""
    student_data['draw_annual'] = ""
    student_data['draw_grade'] = ws["at"+str(row_no)].value
    #PT
    student_data['pt_sem1'] = ""
    student_data['pt_sem2'] = ""
    student_data['pt_annual'] = ""
    student_data['pt_grade'] = ws["au"+str(row_no)].value
    #Other
    student_data['total'] = str(ws["av"+str(row_no)].value)
    student_data['per'] = str(round(ws["aw"+str(row_no)].value,1))
    student_data['grade'] = ws["ax"+str(row_no)].value
    return student_data

pdf = PDF(orientation="L",format="A4")
for no in range(1,no_of_student+1):
    data=fetchData(no)
    result_pdf(data,pdf)
pdf.output("result.pdf")